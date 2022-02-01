from openpyxl import load_workbook

def read(filename):

    wb = load_workbook(filename=filename, read_only=True) 
    worksheets = wb.sheetnames

    for ws in worksheets:
        rows = filter(lambda x: x[1].value != None, wb[ws].rows)

        for row in rows:
            for cell in row:
                print(cell.value, end=' ')
            print()

    wb.close()

def build_request():
    pass
    
if __name__ == '__main__':
    
    try:
        read('./scripts/data.xlsx')
    except:
        pass